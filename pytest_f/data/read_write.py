import openpyxl

class ReadWrite:
    def __init__(self,file,num):
        self.file=file
        self.num=num

    def read(self):
        wb = openpyxl.load_workbook(self.file)
        sheets=wb.sheetnames
        table=wb[sheets[self.num-1]]
        wb.active
        nrows=table.max_row
        ncols=table.max_column
        row_values=[]
        rows=[]
        for row in range(2,nrows+1):
            for col in range(1,ncols+1):
                cell_value=table.cell(row,col).value
                row_values.append(cell_value)
            rows.append(row_values)
            row_values=[]
        return rows
    # [['shelly','p@ssw0rd'],['test','123456']]

    def write(self,username,password):
        wb = openpyxl.load_workbook(self.file)
        sheets = wb.sheetnames
        table = wb[sheets[self.num - 1]]
        table = wb.active
        nrows = table.max_row
        for rows in range(nrows + 1, nrows + 2):
            table.cell(rows,1).value=username
            table.cell(rows,2).value=password
        wb.save(self.file)
        wb.close()